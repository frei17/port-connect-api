#!/usr/bin/env python
# coding: utf-8

# In[1]:


# Import required Python modules and libraries

import requests
import pandas as pd
import openpyxl
import datetime


# In[2]:


# Full access to Port-Connect API port schedule free for Track & Trace subscribed users

api_key = "60f4bc63c0aa4c26ba7e3730401dabd4" 
endpoint = "https://api.portconnect.io/v1/scheduled-vessels?vesselType=COMMERCIAL"


# In[3]:


# Send a request and get a response from the server containing Port-Connect live schedule in JSON format

response = requests.get(endpoint,headers={'Ocp-Apim-Subscription-Key':api_key}).json()

print(response)
# In[4]:


# Create a Dataframe containing Port-Connect live schedule from the server response

port_schedule = pd.DataFrame(response)


# In[5]:


# Read Moveware Emulation Excel spreadsheet

Moveware_emulation = pd.read_excel('Moveware Emulation.xlsx',header=None,na_filter=False)


# In[6]:


# Asssign data from Excel cells to Python variables

vesselName = str(Moveware_emulation.iloc[10,3]).strip()
voyageNumber = str(Moveware_emulation.iloc[11,5]).strip()
portCode = str(Moveware_emulation.iloc[13,3]).strip()
portOfLoading = str(Moveware_emulation.iloc[15,3]).strip()
portOfOrigin = str(Moveware_emulation.iloc[14,3]).strip()


# In[7]:


# Loop through the Dataframe and create a filtered dataframe returning a single row with the lookup vessel, voyage and port code providing there is a match across all three parameters.

port_schedule_filtered = port_schedule[(port_schedule.vesselName.str.contains(vesselName,case=False))&
       (port_schedule.outboundVoyage.str.contains(voyageNumber,case=False))&
       (port_schedule.portCode.str.contains(portCode,case=False))]


# In[8]:


# Assign values from the filtered dataframe to Python variables and covert them to formatted date strings

departureDatetime = pd.Series(pd.to_datetime(port_schedule_filtered.departureDatetime)).dt.strftime('%d/%m/%Y').item()
receivalCommenceInland = pd.Series(pd.to_datetime(port_schedule_filtered.receivalCommenceInland)).dt.strftime('%d-%b-%y %H:%M').item()
receivalCutoffInland = pd.Series(pd.to_datetime(port_schedule_filtered.receivalCutoffInland)).dt.strftime('%d-%b-%y %H:%M').item()
receivalCommenceSeaport = pd.Series(pd.to_datetime(port_schedule_filtered.receivalCommenceSeaport)).dt.strftime('%d-%b-%y %H:%M').item()
receivalCutoffSeaport = pd.Series(pd.to_datetime(port_schedule_filtered.receivalCutoffSeaport)).dt.strftime('%d-%b-%y %H:%M').item()


# In[9]:


# Define functions for each port that write text and dates contained in Python variables to the Excel spreadsheet

def inland():
    wb = openpyxl.load_workbook('Moveware Emulation.xlsx')
    wb['EX Waybill']['D13'] = departureDatetime
    wb['EX Waybill']['V6'] = "Metro "+receivalCommenceInland+" - "+receivalCutoffInland
    wb['EX Waybill']['P6'] = datetime.date.today()
    wb.save('Moveware Emulation.xlsx')    


# In[10]:


def seaport_TRG():
    wb = openpyxl.load_workbook('Moveware Emulation.xlsx')
    wb['EX Waybill']['D13'] = departureDatetime
    wb['EX Waybill']['V6'] = "Port cut-off: "+receivalCutoffSeaport
    wb['EX Waybill']['P6'] = datetime.date.today()
    wb.save('Moveware Emulation.xlsx')


# In[11]:


def seaport_AKL():
    wb = openpyxl.load_workbook('Moveware Emulation.xlsx')
    wb['EX Waybill']['D13'] = departureDatetime
    wb['EX Waybill']['V6'] = "POAL cut-off: "+receivalCutoffSeaport
    wb['EX Waybill']['P6'] = datetime.date.today()
    wb.save('Moveware Emulation.xlsx')


# In[12]:


def seaport_TIMARU():
    wb = openpyxl.load_workbook('Moveware Emulation.xlsx')
    wb['EX Waybill']['D13'] = departureDatetime
    wb['EX Waybill']['V6'] = "Port cut-off: "+receivalCutoffSeaport
    wb['EX Waybill']['P6'] = datetime.date.today()
    wb.save('Moveware Emulation.xlsx')


# In[13]:


# Specify conditions that invoke an appropriate function based on selected port

if portCode.upper() == "NZTRG":
    if "Metroport" in portOfOrigin:
        inland()
    else:
        seaport_TRG()
elif portCode.upper() == "NZAKL":
    seaport_AKL()
else:
    seaport_TIMARU()

